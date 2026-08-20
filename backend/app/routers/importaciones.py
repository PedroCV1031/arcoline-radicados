from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import numpy as np
import openpyxl
import pandas as pd
from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from app.database import client, radicados_collection

router = APIRouter(
    prefix="/api/importaciones",
    tags=["Importaciones"]
)

EXTENSIONES_PERMITIDAS = {".xlsx", ".xlsm"}
TAMANO_MAXIMO = 20 * 1024 * 1024

COLUMNAS_ESPERADAS = [
    "Fecha inicio",
    "Fecha limite",
    "Cliente",
    "Orden de compra",
    "Referencia",
    "Talla",
    "Cantidad",
    "Unidades despachadas",
    "Fecha entrega final",
]


def es_hoja_valida(nombre_hoja: str) -> bool:
    nombre_normalizado = nombre_hoja.strip().lower()

    return not nombre_normalizado.startswith(
        ("hoja", "sheet")
    )


def validar_archivo(
    nombre_archivo: str,
    contenido: bytes
) -> None:
    extension = Path(nombre_archivo).suffix.lower()

    if extension not in EXTENSIONES_PERMITIDAS:
        raise HTTPException(
            status_code=400,
            detail="Solo se permiten archivos .xlsx o .xlsm"
        )

    if not contenido:
        raise HTTPException(
            status_code=400,
            detail="El archivo está vacío"
        )

    if len(contenido) > TAMANO_MAXIMO:
        raise HTTPException(
            status_code=413,
            detail="El archivo supera el límite de 20 MB"
        )


def normalizar_valor(valor):
    if valor is None:
        return None

    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()

    if isinstance(valor, np.generic):
        return valor.item()

    if isinstance(valor, str):
        valor = valor.strip()
        return valor if valor else None

    return valor


def convertir_identificador_a_texto(valor):
    valor = normalizar_valor(valor)

    if valor is None:
        return None

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()


def procesar_hoja(
    contenido: bytes,
    nombre_archivo: str,
    nombre_hoja: str,
    lote_importacion: str
) -> list[dict]:
    try:
        dataframe = pd.read_excel(
            BytesIO(contenido),
            sheet_name=nombre_hoja,
            engine="openpyxl"
        )
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                f"No fue posible procesar la hoja "
                f"'{nombre_hoja}': {error}"
            )
        )

    if dataframe.empty:
        raise HTTPException(
            status_code=400,
            detail=f"La hoja '{nombre_hoja}' está vacía"
        )

    dataframe.columns = [
        str(columna).strip()
        for columna in dataframe.columns
    ]

    columnas_faltantes = [
        columna
        for columna in COLUMNAS_ESPERADAS
        if columna not in dataframe.columns
    ]

    if columnas_faltantes:
        raise HTTPException(
            status_code=400,
            detail={
                "mensaje": (
                    f"La hoja '{nombre_hoja}' no tiene "
                    "la estructura esperada"
                ),
                "columnas_faltantes": columnas_faltantes
            }
        )

    dataframe = dataframe[COLUMNAS_ESPERADAS]
    dataframe = dataframe.dropna(how="all")

    if dataframe.empty:
        raise HTTPException(
            status_code=400,
            detail=(
                f"La hoja '{nombre_hoja}' no contiene "
                "registros para importar"
            )
        )

    registros = []
    fecha_importacion = datetime.now(timezone.utc)

    for indice, fila in dataframe.iterrows():
        registro = {
            columna: normalizar_valor(fila[columna])
            for columna in COLUMNAS_ESPERADAS
        }

        registro["Cliente"] = convertir_identificador_a_texto(
            registro["Cliente"]
        )

        registro["Orden de compra"] = convertir_identificador_a_texto(
            registro["Orden de compra"]
        )

        registro["Referencia"] = convertir_identificador_a_texto(
            registro["Referencia"]
        )

        registro["_metadatos"] = {
            "archivo_origen": nombre_archivo,
            "hoja_origen": nombre_hoja,
            "fila_origen": int(indice) + 2,
            "fecha_importacion": fecha_importacion,
            "lote_importacion": lote_importacion
        }

        registros.append(registro)

    return registros


@router.post("/hojas")
async def consultar_hojas(
    archivo: UploadFile = File(...)
):
    nombre_archivo = Path(archivo.filename or "").name
    contenido = await archivo.read()

    validar_archivo(nombre_archivo, contenido)

    libro = None

    try:
        libro = openpyxl.load_workbook(
            BytesIO(contenido),
            read_only=True,
            data_only=True
        )

        todas_las_hojas = libro.sheetnames

        hojas_validas = [
            hoja
            for hoja in todas_las_hojas
            if es_hoja_valida(hoja)
        ]

        hojas_ignoradas = [
            hoja
            for hoja in todas_las_hojas
            if not es_hoja_valida(hoja)
        ]

        if not hojas_validas:
            raise HTTPException(
                status_code=400,
                detail="El archivo no contiene hojas válidas"
            )

        return {
            "archivo": nombre_archivo,
            "total_hojas": len(todas_las_hojas),
            "hojas_validas": hojas_validas,
            "hojas_ignoradas": hojas_ignoradas
        }

    except HTTPException:
        raise

    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=f"No fue posible leer el archivo: {error}"
        )

    finally:
        if libro is not None:
            libro.close()


@router.post("/cargar")
async def cargar_hojas(
    archivo: UploadFile = File(...),
    hojas: list[str] = Form(...),
    confirmar: bool = Form(False)
):
    if not confirmar:
        raise HTTPException(
            status_code=400,
            detail="Debe confirmar el reemplazo de las hojas"
        )

    nombre_archivo = Path(archivo.filename or "").name
    contenido = await archivo.read()

    validar_archivo(nombre_archivo, contenido)
    
    hojas_separadas = []

    for valor in hojas:
        hojas_separadas.extend(
            hoja.strip()
            for hoja in valor.split(",")
            if hoja.strip()
        )

    hojas_seleccionadas = list(
        dict.fromkeys(hojas_separadas)
    )

    if not hojas_seleccionadas:
        raise HTTPException(
            status_code=400,
            detail="Debe seleccionar al menos una hoja"
        )

    libro = None

    try:
        libro = openpyxl.load_workbook(
            BytesIO(contenido),
            read_only=True,
            data_only=True
        )

        hojas_disponibles = libro.sheetnames

    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=f"No fue posible abrir el archivo: {error}"
        )

    finally:
        if libro is not None:
            libro.close()

    hojas_invalidas = [
        hoja
        for hoja in hojas_seleccionadas
        if (
            hoja not in hojas_disponibles
            or not es_hoja_valida(hoja)
        )
    ]

    if hojas_invalidas:
        raise HTTPException(
            status_code=400,
            detail={
                "mensaje": "Hay hojas inexistentes o no permitidas",
                "hojas_invalidas": hojas_invalidas
            }
        )

    lote_importacion = str(uuid4())
    registros_por_hoja = {}

    # Primero procesa y valida todo, sin modificar MongoDB.
    for hoja in hojas_seleccionadas:
        registros_por_hoja[hoja] = procesar_hoja(
            contenido=contenido,
            nombre_archivo=nombre_archivo,
            nombre_hoja=hoja,
            lote_importacion=lote_importacion
        )

    resultado_hojas = []

    try:
        # La transacción garantiza que el reemplazo sea completo.
        with client.start_session() as sesion:
            with sesion.start_transaction():
                for hoja, registros in registros_por_hoja.items():
                    eliminacion = radicados_collection.delete_many(
                        {
                            "_metadatos.hoja_origen": hoja
                        },
                        session=sesion
                    )

                    insercion = radicados_collection.insert_many(
                        registros,
                        session=sesion
                    )

                    resultado_hojas.append({
                        "hoja": hoja,
                        "registros_eliminados": (
                            eliminacion.deleted_count
                        ),
                        "registros_insertados": len(
                            insercion.inserted_ids
                        )
                    })

    except Exception as error:
        raise HTTPException(
            status_code=500,
            detail=(
                "No fue posible reemplazar las hojas en "
                f"MongoDB: {error}"
            )
        )

    return {
        "mensaje": "Importación completada correctamente",
        "archivo": nombre_archivo,
        "lote_importacion": lote_importacion,
        "hojas_procesadas": resultado_hojas,
        "total_insertados": sum(
            resultado["registros_insertados"]
            for resultado in resultado_hojas
        ),
        "total_eliminados": sum(
            resultado["registros_eliminados"]
            for resultado in resultado_hojas
        )
    }